from typing import Dict

from openpyxl.styles import Alignment
from openpyxl.workbook import Workbook


def _format_cells_in_xlsx(ws: Workbook) -> None:
    """Выравнивает текст в ячейках и задает ширину столбцам по максимальному значению длины в ячейках
    Не стал разделять на две отдельные функции, так как два раза проходилсь бы по тому же вложенному циклу"""

    # Создаем объект стиля с выравниванием по центру
    center_alignment = Alignment(horizontal='center', vertical='center')

    for column in ws.columns:
        max_length = 0
        column_name = column[0].column_letter  # Получаем буквенное обозначение столбца (для установления ширины)

        for cell in column:
            cell.alignment = center_alignment  # Применяем стиль выравнивания к каждому столбцу

            # Находим максимальную длину значения в столбце (для установления ширины)
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except Exception:
                pass

        # Устанавливаем ширину столбца на основе максимальной длины значения + дополнительное значение
        adjusted_width = (max_length + 2)
        ws.column_dimensions[column_name].width = adjusted_width


def create_week_repost_as_xlsx_file(models: list, prepared_pages_data: Dict[str, list], timestamp: str) -> Workbook:
    """ создаем xlsx файл с отчетом по каждой модели и серии робатов"""

    # создаем наш xlsx файл и разбиваем на отдельные листы
    wb = Workbook()
    for num, model in enumerate(models):
        wb.create_sheet(model['name'], num)
    del wb[wb.sheetnames[-1]]  # удаляем изначально созданный лист в xlsx файле

    # заполняем страницы построчно данными из prepared_pages_data
    for model in models:
        ws = wb[model['name']]
        for raw in prepared_pages_data[model['name']]:
            ws.append(raw)
            _format_cells_in_xlsx(ws=ws)
    return wb